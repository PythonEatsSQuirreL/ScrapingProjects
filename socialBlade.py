#importing the libraries
import re
import time
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common import keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import os,sys
import openpyxl
import xlsxwriter
import io
import csv
from csv import writer
import getpass
from time import sleep
import pandas as pd
import  tkinter as tk
from tkinter import filedialog
from tkinter import simpledialog
import pyperclip as pc
import shutil
import requests
import win32api
import os.path




#pc's user
user = os.environ.get('USERNAME')
file_name = str(os.path.basename(sys.argv[0]))

#file path
now = datetime.now()
path = os.path.abspath(".")
runPath = path
usrName = getpass.getuser()

print("Automation started!")
#creating a Chrome profile folder in order to remain logged in on the SocialBlade platform
def chromeProfileFolder(username):
    toolName = os.path.basename(__file__)
    try:
        toolName = toolName.split(".py")[0]
    except:
        toolName = toolName.split(".exe")[0]
    if " " in toolName:
        profileName = toolName.split(" ")[0]
    else:
        profileName = toolName

    # Checking for Chrome Profiles folder
    ChromeProfileDir = ("Chrome Profiles")
    profilesPathToCheck = ("C:\\Users\\" + username + "\\" + ChromeProfileDir)
    CheckProf = os.path.isdir(profilesPathToCheck)
    if not CheckProf:
        os.makedirs(profilesPathToCheck)

    # End of Checking for Chrome Profiles folder

    # Checking for automation profile
    MYDIR = profileName
    Prof_path = ("C:\\Users\\" + username + "\\Chrome Profiles\\" + MYDIR)
    CHECK_FOLDER = os.path.isdir(Prof_path)
    print(str(Prof_path))
    if not CHECK_FOLDER:
        #os.makedirs(Prof_path)
        print("Chrome Profile not found for this automation.\nCreating Chrome Profile...\nPlease wait!")
        src = "C:\\Users\\" + username + "\\AppData\\Local\\Google\\Chrome\\User Data"
        src2 = "D:\\Users\\" + username + "\\AppData\\Local\\Google\\Chrome\\User Data"
        if os.path.isdir(src):
            shutil.copytree(src, Prof_path)
        elif os.path.isdir(src2):
            shutil.copytree(src2, Prof_path)
        print("Chrome Profile created!")
    # End of Checking for automation profile
    return Prof_path
#check if we have internet connection, if not, wait until we will get some connection
def internetConnection ():
    url = "https://www.google.com/"
    timeout = 5
    try:
        request = requests.get(url, timeout = timeout)
        connected = "connected"
    except (requests. ConnectionError, requests. Timeout) as exception:
        connected = "disconnected"
    
    if connected == "disconnected":
        print("It seems the internet is down! Please wait for the connection to be restored...")
        while connected == "disconnected":
            try:
                request = requests.get(url, timeout = timeout)
                connected = "connected"
            except (requests. ConnectionError, requests. Timeout) as exception:
                connected = "disconnected"
        print("Internet restored! Resuming...")

chrPath = chromeProfileFolder(usrName)
os.environ['WDM_LOG_LEVEL'] = '0'
options = Options()
options.headless = False
options.add_argument("--log-level=3")
options.add_argument("--disable-logging")
prefs = {'download.default_directory' : runPath}
options.add_experimental_option('prefs', prefs)
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument("user-data-dir=" + chrPath)
selenium = webdriver.Chrome(ChromeDriverManager().install(), options=options)
selenium.maximize_window()
internetConnection ()
#initialize the backup
backup = False
#initialize the saved which will be incremented later
saved = 0
#define the excel workbook
file_exists = os.path.exists('youtubeCheck.xlsx')

if file_exists:
    ex = "youtubeCheck.xlsx"
else:
    workbook = xlsxwriter.Workbook('youtubeCheck.xlsx')
    worksheet = workbook.add_worksheet('Sheet1')
    workbook.close()
    ex = "youtubeCheck.xlsx"
#open the workbook
videosDB = pd.read_excel(ex, sheet_name="Sheet1")

wbk = openpyxl.load_workbook(str(ex))

sh = wbk.active
#wait ultil everything loads
WebDriverWait(selenium, 2).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
wait = WebDriverWait(selenium, 60, poll_frequency=1)
#open a new selenium tab
selenium.execute_script("window.open('about:blank','firsttab');")
selenium.switch_to.window("firsttab")
#get the SocialBlade link that you want to check
selenium.get('https://socialblade.com/youtube/user/rockatalynutz')
WebDriverWait(selenium, 2).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
selenium.implicitly_wait(1)
#open a new selenium tab
selenium.execute_script("window.open('about:blank','secondtab');")
selenium.switch_to.window("secondtab")
#get the second SocialBlade link that you want to check
selenium.get('https://socialblade.com/youtube/channel/UCnU1hpSB-hCqo4DdFfsLfuQ')
WebDriverWait(selenium, 2).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
selenium.implicitly_wait(1)
#open a new selenium tab
selenium.execute_script("window.open('about:blank','thirdtab');")
selenium.switch_to.window("thirdtab")
#get the third SocialBlade link that you want to check
selenium.get('https://socialblade.com/instagram/user/pythoneatssquirrel')
WebDriverWait(selenium, 2).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
selenium.implicitly_wait(1)

selenium.switch_to.window("firsttab")
wait = WebDriverWait(selenium, 60, poll_frequency=1)
#wait 2 seconds in order for everything to load
time.sleep(2)
#define ii in order to increment it later
ii = 1
#create a results array
results = []
#get the firsth path for the element that we need
uploads1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[2]/span[2]') #get title column
#extract only the text from it
titl = uploads1.text
#print it
print("str titl: " + str(titl))
#add it to the excel file
keywords_val = sh.cell(row=ii, column=1)
#add it to the DB
results.append(titl)
print("Added " + titl + " to the list")
#define the new row number
row = len(videosDB)
print("Row: " + str(row + 1))
#define the time and date when the automation runned
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
#add it to the DB
videosDB.loc[row, "RandomFunUploads"] = titl
#increment the saved
saved += 1
#the same steps will be repeated for every row that we want to add, and every data that we need to scrape from SocialBlade
subscribers1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[3]/span[2]') #get title column
titl = subscribers1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "RandomFunSubscribers"] = titl
saved += 1

videoviews1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[4]/span[2]') #get title column
titl = videoviews1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)    
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "RandomFunVideo views"] = titl
saved += 1

timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "RandomFunDate check"] = timestampstart
saved += 1

selenium.switch_to.window("secondtab")
WebDriverWait(selenium, 3).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
selenium.implicitly_wait(1)
time.sleep(4)

uploads1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[2]/span[2]') #get title column
titl = uploads1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "PythonEatsSQuirrelUploads"] = titl
saved += 1

subscribers1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[3]/span[2]') #get title column
titl = subscribers1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "PythonEatsSQuirrelSubscribers"] = titl
saved += 1

videoviews1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[4]/span[2]') #get title column
titl = videoviews1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "PythonEatsSQuirrelVideo views"] = titl
saved += 1

videosDB.loc[row, "PythonEatsSQuirrelDate check"] = timestampstart
saved += 1

selenium.switch_to.window("thirdtab")
WebDriverWait(selenium, 2).until(lambda driver: selenium.execute_script('return document.readyState') == 'complete')
selenium.implicitly_wait(1)
time.sleep(2)

mediauploads1 = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[2]/span[2]') #get title column
titl = mediauploads1.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "InstaUploads"] = titl
saved += 1

instafolowers = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[3]/span[2]') #get title column
titl = instafolowers.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "InstaFollowers"] = titl
saved += 1

instarate = selenium.find_element_by_xpath('//*[@id="YouTubeUserTopInfoBlock"]/div[5]/span[3]') #get title column
titl = instarate.text
print("Video: " + str(titl))         
keywords_val = sh.cell(row=ii, column=1)   
results.append(titl)
print("Added " + titl + " to the list") 
print("Row: " + str(row + 1))
timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "InstaRate"] = titl
saved += 1

timestampstart = time.strftime("%Y_%m_%d-%H:%M:%S")
videosDB.loc[row, "InstaDate"] = timestampstart
saved += 1
#save the data into the DB once every 5 datas
if saved >= 5:
     filename3 = ('youtubeCheck.xlsx')
     videosDB.to_excel(filename3,sheet_name='Sheet1',header=True, index=False)
     saved = 0
     backup = True
#create a backup file
if backup == True:
     shutil.copy(filename3, "youtubeCheckBackup.xlsx")
     backup = False

time.sleep(2)
selenium.switch_to.window("firsttab")
#save again the DB to the excel file in case anything happend before
videosDB.to_excel(ex,sheet_name='Sheet1',header=True, index=False)
#quit the selenium instance
selenium.quit()
#open Tkinter in order to show a pop-up notification
root = tk.Tk()
root.withdraw()
#show the pop-up above everything else
root.wm_attributes('-topmost', 1)
#define the message
tk.messagebox.showinfo(title=None, message="Scraping completed!")
